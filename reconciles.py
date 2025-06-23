import os
import pandas as pd
from collections import defaultdict
import pathlib
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def mark_items_with_colors(file_path, sheet_names, inconsistent_names, unique_names, consistent_names, name_col="品名", qty_col="本次领用", diff_dict=None):
    """
    标记数目不一致为荧光色，数目一致为绿色，只在本表有的品名为红色
    """
    fill_yellow = PatternFill(fill_type="solid", fgColor="FFFF00")  # 荧光色
    fill_green = PatternFill(fill_type="solid", fgColor="ADD88D")   # 绿色
    fill_red = PatternFill(fill_type="solid", fgColor="FF0000")     # 红色

    wb = openpyxl.load_workbook(file_path)
    for sheet in sheet_names:
        ws = wb[sheet]
        # 找到表头行和列
        header_row = None
        name_col_idx = None
        qty_col_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5), 1):
            for cell in row:
                if cell.value == name_col:
                    header_row = i
                    name_col_idx = cell.col_idx
                if cell.value == qty_col:
                    qty_col_idx = cell.col_idx
            if header_row and name_col_idx and qty_col_idx:
                break
        if not header_row or not name_col_idx or not qty_col_idx:
            continue
        print(f"处理表: {sheet}, 表头行: {header_row}, 品名列: {name_col_idx}, 数量列: {qty_col_idx}")
    
        # 标色
        for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
            name_cell = row[name_col_idx - 1]
            qty_cell = row[qty_col_idx - 1]
            item_name = str(name_cell.value).strip() if name_cell.value else ""
            if item_name in inconsistent_names:
                name_cell.fill = fill_yellow
                qty_cell.fill = fill_yellow
            elif item_name in consistent_names:
                name_cell.fill = fill_green
                qty_cell.fill = fill_green
            elif item_name in unique_names:
                name_cell.fill = fill_red
             
        # 在表头右侧添加图注
        legend_start_col = ws.max_column + 2
        ws.cell(row=header_row, column=legend_start_col).fill = fill_yellow
        ws.cell(row=header_row, column=legend_start_col+1).fill = fill_green
        ws.cell(row=header_row, column=legend_start_col+2).fill = fill_red
        ws.cell(row=header_row+1, column=legend_start_col, value="数目不一致")
        ws.cell(row=header_row+1, column=legend_start_col+1, value="数目一致")
        ws.cell(row=header_row+1, column=legend_start_col+2, value="只在本表有的品名")
        # 设置列宽
        for i in range(3):
            col_letter = get_column_letter(legend_start_col + i)
            ws.column_dimensions[col_letter].width = 15

        # 差异目录插入
        if diff_dict:
            max_row = ws.max_row
            ws.cell(row=max_row+2, column=1, value="差异目录")
            ws.cell(row=max_row+3, column=1, value="品名")
            ws.cell(row=max_row+3, column=2, value="领用单数量")
            ws.cell(row=max_row+3, column=3, value="盘点表数量")
            r = max_row+4
            for item in diff_dict:
                ws.cell(row=r, column=1, value=item['品名'])
                ws.cell(row=r, column=2, value=item['领用单数量'])
                ws.cell(row=r, column=3, value=item['盘点数量'])
                r += 1

    # 输出到同目录，文件名前加“标色_”
    out_path = os.path.join(os.path.dirname(file_path), f"标色_{os.path.basename(file_path)}")
    wb.save(out_path)
    print(f"已输出标色文件: {out_path}")

def process_inventory_data(file_path, sheet_names):
    """处理盘点数据，汇总品名的本次领用数量"""
    inventory_data = defaultdict(int)
    try:
        for sheet in sheet_names:
            try:
                # 强制第二行为表头
                df = pd.read_excel(file_path, sheet_name=sheet, header=1)
                print(f"{file_path} - {sheet} 表头: {df.columns.tolist()}")
                print(df.head())

                for _, row in df.iterrows():
                    item_name = str(row.get("品名", "")).strip()
                    usage = row.get("本次领用", None)
                    if item_name == "" or item_name == "nan" or pd.isna(usage):
                        continue
                    if pd.api.types.is_number(usage):
                        inventory_data[item_name] += int(usage)
                        # print(f"累计盘点: {item_name} += {usage}")
            except Exception as e:
                print(f"处理盘点表 {sheet} 时出错: {str(e)}")
                continue
            
    except Exception as e:
        print(f"读取Excel文件时出错: {str(e)}")
    return dict(inventory_data)

def process_requisition_data(file_path, sheet_name):
    requisition_data = defaultdict(int)
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=2)
        print(f"{file_path} 领用单表头: {df.columns.tolist()}")
        print(df.head())

        item_col = '品名'
        qty_col = '数量'
        
        for _, row in df.iterrows():
            item_name = str(row[item_col]).strip()
            qty = row[qty_col]
            # 过滤掉无效品名和数量
            if pd.isna(item_name) or item_name == "" or item_name.lower() == "nan" or pd.isna(qty):
                continue
            if pd.api.types.is_number(qty):
                requisition_data[item_name] += int(qty)
                # print(f"累计领用: {item_name} += {qty}")
    except Exception as e:
        print(f"处理领用单时出错: {str(e)}")
    return dict(requisition_data)

def compare_data(inventory_data, requisition_data):
    """比较盘点和领用单品名的数量"""
    consistent_items = []
    inconsistent_items = []
    
    # 只取交集
    common_items = set(inventory_data.keys()) & set(requisition_data.keys())
    only_in_inventory = set(inventory_data.keys()) - set(requisition_data.keys())
    only_in_requisition = set(requisition_data.keys()) - set(inventory_data.keys())

    for item in common_items:
        inv_qty = inventory_data.get(item, 0)
        req_qty = requisition_data.get(item, 0)
        
        if inv_qty == req_qty:
            consistent_items.append((item, inv_qty))
        else:
            inconsistent_items.append({
                '品名': item,
                '盘点数量': inv_qty,
                '领用单数量': req_qty,
                '差异': abs(inv_qty - req_qty)
            })
    
    return consistent_items, inconsistent_items, only_in_inventory, only_in_requisition

def print_items_per_line(items, per_line=10):
    items = list(items)
    for i in range(0, len(items), per_line):
        print("，".join(items[i:i+per_line]))

def main():
    # 只需选择一个Excel文件
    project_dir = pathlib.Path(__file__).parent
    excel_path = project_dir / "盘点" / "五金盘点表.xlsx"
    # 盘点sheet名
    inventory_sheets = ["电类盘点", "水类盘点"]
    requisition_sheet = "领用单"
    
    # 处理盘点数据
    all_inventory = process_inventory_data(excel_path, inventory_sheets)
    # 处理领用单数据
    all_requisition = process_requisition_data(excel_path, requisition_sheet)
    
    # 比较数据
    consistent, inconsistent, only_in_inventory, only_in_requisition = compare_data(all_inventory, all_requisition)
    
    print("盘点汇总：", all_inventory)
    print("领用汇总：", all_requisition)

    # 输出结果
    print("\n一致的品名及数量:")
    for item, qty in consistent:
        print(f"{item}: {qty}")
    
    print("\n不一致的品名:")
    print("品名 | 盘点数量 | 领用单数量 | 差异")
    for detail in inconsistent:
        print(f"{detail['品名']} | {detail['盘点数量']} | {detail['领用单数量']} | {detail['差异']}")

    print("\n只在盘点表有的品名（请核查）:")
    if only_in_inventory:
        print_items_per_line(sorted(only_in_inventory), per_line=10)
    else:
        print("无")

    print("\n只在领用单有的品名（请核查）:")
    if only_in_requisition:
        print_items_per_line(sorted(only_in_requisition), per_line=10)
    else:
        print("无")

    # 标记不一致的品名
    if consistent or inconsistent or only_in_inventory:
        mark_items_with_colors(
            excel_path,
            sheet_names=inventory_sheets,
            inconsistent_names=set(item['品名'] for item in inconsistent),
            unique_names=only_in_inventory,
            consistent_names=set(item for item, _ in consistent),
            name_col="品名",
            qty_col="本次领用",
            diff_dict=[
                {'品名': d['品名'], '领用单数量': d['领用单数量'], '盘点数量': d['盘点数量']}
                for d in inconsistent
            ]
        )
    if consistent or inconsistent or only_in_requisition:
        mark_items_with_colors(
            excel_path,
            sheet_names=[requisition_sheet],
            inconsistent_names=set(item['品名'] for item in inconsistent),
            unique_names=only_in_requisition,
            consistent_names=set(item for item, _ in consistent),
            name_col="品名",
            qty_col="数量",
            diff_dict=[
                {'品名': d['品名'], '领用单数量': d['领用单数量'], '盘点数量': d['盘点数量']}
                for d in inconsistent
            ]
        )

if __name__ == "__main__":
    main()